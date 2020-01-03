import os
import openpyxl

class GSV:
    def gsv(player_id, status, value):
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot/player_management')
        file1 = openpyxl.load_workbook('status.xlsx')
        sheet1 = file1.active
        idx1 = 2
        if status != sheet1.cell(idx1, 1).value:
            idx1 += 1
        idx2 = 2
        if player_id != sheet1.cell(1, idx2).value:
            idx2 += 1
        if sheet1.cell(idx1, idx2).value == None:
            sheet1.cell(idx1, idx2).value = 0
        sheet1.cell(idx1, idx2).value = sheet1.cell(idx1, idx2).value + int(value)
        file1.save("status.xlsx")
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot')
        return
