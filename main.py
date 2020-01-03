import discord
import openpyxl
import os
from game_system import *

client = discord.Client()

os.chdir('C:/Users/airsk/Desktop/TRPG_Bot/player_management')
file1 = openpyxl.load_workbook("status.xlsx")
status_sheet = file1.active
file2 = openpyxl.load_workbook("get_item.xlsx")
get_item_sheet = file2.active
os.chdir('C:/Users/airsk/Desktop/TRPG_Bot')

@client.event
async def on_ready():
    print("ready")
    game = discord.Game("/ 명령어")
    await client.change_presence(status=discord.Status.online, activity=game)

@client.event
async def on_message(message):

    if message.author.bot:
        return None
    id = message.author.id

    if message.content.startswith("/ 명령어"):
        embed = discord.Embed(title="명령어", colour=discord.Colour.blue())
        embed.add_field(name="캐릭터 생성", value="/ 캐릭터 생성", inline=True)
        embed.add_field(name="스탯 확인", value="/ 스탯 확인 @이름", inline=True)
        embed.add_field(name="스탯 부여", value="/ 부여 @이름", inline=True)
        embed.add_field(name="아이템 지급", value="/ 지급 @이름", inline=True)
        embed.add_field(name="인벤토리 확인", value="/ 인벤토리 @이름", inline=True)
        embed.add_field(name="대상에게 공격", value="/ 공격 @이름", inline=True)
        embed.add_field(name="주사위 굴리기", value="/ 2d6 / d눈금수", inline=True)
        await message.channel.send(embed=embed)

    if message.content.startswith("/ 캐릭터 생성"):
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot/player_management')
        if id in status_sheet[1]:
            return None
        else:
            idx1 = 1
            while status_sheet.cell(1, idx1).value != None:
                idx1 += 1
            status_sheet.cell(1, idx1).value = str(id)
            status_sheet.cell(2, idx1).value = 9
            status_sheet.cell(3, idx1).value = 0
            status_sheet.cell(4, idx1).value = 0
            status_sheet.cell(5, idx1).value = 0
            status_sheet.cell(6, idx1).value = 0
            status_sheet.cell(7, idx1).value = 0
            status_sheet.cell(8, idx1).value = 0
            file1.save("status.xlsx")

            idx2 = 1
            while get_item_sheet.cell(1, idx2).value != None:
                idx2 += 1
            get_item_sheet.cell(1, idx2).value = str(id)
            get_item_sheet.cell(1, (idx2 + 1)).value = str('number')
            file2.save("get_item.xlsx")
            await message.channel.send("캐릭터가 생성되었습니다!")
            os.chdir('C:/Users/airsk/Desktop/TRPG_Bot')

    if message.content.startswith("/ 스탯 확인"):
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot/player_management')
        player_id = message.content[10:28]
        idx = 1
        while player_id != status_sheet.cell(1, idx).value:
            idx += 1
        embed = discord.Embed(colour=discord.Colour.blue())
        embed.add_field(name="플레이어의 이름", value=("<@%s>" %str(status_sheet.cell(1, idx).value)), inline=True)
        embed.add_field(name="플레이어의 직업", value=str("제작중"), inline=True)
        embed.add_field(name="플레이어의 현재 HP", value=str(status_sheet.cell(2, idx).value), inline=True)
        embed.add_field(name="근력", value=str(status_sheet.cell(3, idx).value), inline=True)
        embed.add_field(name="민첩성", value=str(status_sheet.cell(4, idx).value), inline=True)
        embed.add_field(name="체력", value=str(status_sheet.cell(5, idx).value), inline=True)
        embed.add_field(name="지능", value=str(status_sheet.cell(6, idx).value), inline=True)
        embed.add_field(name="지혜", value=str(status_sheet.cell(7, idx).value), inline=True)
        embed.add_field(name="매력", value=str(status_sheet.cell(8, idx).value), inline=True)
        embed.set_thumbnail(url=message.author.avatar_url)
        await message.channel.send(embed=embed)
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot')

    if message.content.startswith("/ 부여"):
        player_id = message.content[7:25]
        status = message.content[27:30]
        value = message.content[31:]
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot/player_management')
        idx1 = 2
        while status != status_sheet.cell(idx1, 1).value:
            idx1 += 1
        idx2 = 2
        while player_id != status_sheet.cell(1, idx2).value:
            idx2 += 1
        if status_sheet.cell(idx1, idx2).value == None:
            status_sheet.cell(idx1, idx2).value = 0
        status_sheet.cell(idx1, idx2).value = status_sheet.cell(idx1, idx2).value + int(value)
        file1.save("status.xlsx")
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot')
        await message.channel.send("스탯을 부여하였습니다!")

    if message.content.startswith("/ 지급"):
        order_split = message.content.split()
        player_id = order_split[2]
        player_id = player_id[2:20]
        item_name = order_split[3]
        order_value = order_split[4]
        order_value = order_value.split("개")
        value = order_value[0]

        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot/player_management')
        idx2 = 2
        while player_id != get_item_sheet.cell(1, idx2).value:
            idx2 += 1
        idx1 = 1
        while get_item_sheet.cell(idx1, idx2).value != item_name:
            idx1 += 1
            if get_item_sheet.cell(idx1, idx2).value == None:
                break
        if get_item_sheet.cell(idx1, idx2).value == item_name:
            get_item_sheet.cell(idx1, (idx2 + 1)).value = get_item_sheet.cell(idx1, (idx2 + 1)).value + int(value)
        if get_item_sheet.cell(idx1, idx2).value == None:
            get_item_sheet.cell(idx1, idx2).value = item_name
            get_item_sheet.cell(idx1, (idx2 + 1)).value = 0
            get_item_sheet.cell(idx1, (idx2 + 1)).value = get_item_sheet.cell(idx1, (idx2 + 1)).value + int(value)
        file2.save("get_item.xlsx")
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot')
        await message.channel.send("지급하였습니다!")

    if message.content.startswith("/ 인벤토리"):
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot/player_management')
        embed = discord.Embed(title="인벤토리", colour=discord.Colour.purple())
        player_id = message.content[9:27]
        idx2 = 1
        while player_id != get_item_sheet.cell(1, idx2).value:
            idx2 += 1
        idx1 = 2
        while get_item_sheet.cell(idx1, idx2).value != None:
            value = get_item_sheet.cell(idx1, (idx2 + 1)).value
            embed.add_field(name=get_item_sheet.cell(idx1, idx2).value, value=str(value), inline=True)
            idx1 += 1
        await message.channel.send(embed=embed)
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot')

    if message.content.startswith("/ 공격"):
        embed = discord.Embed(title="전투 내역", colour=discord.Colour.red())
        damage = Dice_make.dice_make(6)
        player_id = message.content[8:26] #공력을 당하는 플레이어의 id
        print(player_id)
        record = ("<@%s>에게 %d의 데미지" %(player_id, damage))
        attack_nick = ("<@%s>" %message.author.id)
        defence_nick = ("<@%s>" %player_id)
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot/player_management')
        embed.add_field(name="공격하는 대상 ", value=attack_nick, inline=True)
        embed.add_field(name="공격 받는 대상 ", value=defence_nick, inline=True)
        embed.add_field(name="대상이 받은 데미지", value=damage, inline=True)
        idx = 1
        while status_sheet.cell(1, idx).value != player_id:
            idx += 1
        remaining_hp = status_sheet.cell(2, idx).value - damage
        status_sheet.cell(2, idx).value = remaining_hp
        embed.add_field(name="남은 체력", value=remaining_hp, inline=True)
        embed.add_field(name="전투 기록", value=record, inline=True)
        await message.channel.send(embed=embed)
        file1.save("status.xlsx")
        os.chdir('C:/Users/airsk/Desktop/TRPG_Bot')

    if message.content.startswith("/ d"):
        print("ok")
        num = message.content[3:]
        number = int(num)
        print(number)
        dice = Dice_make.dice_make(number)
        embed = discord.Embed(title="결과", colour=discord.Colour.green())
        embed.add_field(name="주사위 결과", value=dice)
        await message.channel.send(embed=embed)

    if message.content.startswith("/ 2d6"):
        dice1 = Dice_make.dice_make(6)
        dice2 = Dice_make.dice_make(6)
        event_result = Dice_make.dice_result(dice1, dice2)
        dice_result = ("[%d] [%d]" % (dice1, dice2))
        embed = discord.Embed(title="결과", colour=discord.Colour.green())
        embed.add_field(name="주사위 결과", value=dice_result)

        if event_result == 0:
            embed.add_field(name="이벤트 결과", value="실패")
        if event_result == 1:
            embed.add_field(name="이벤트 결과", value="애매한 성공")
        if event_result == 2:
            embed.add_field(name="이벤트 결과", value="성공")
        await message.channel.send(embed=embed)

client.run('NjU2OTAwMTY2OTUwOTc3NTM3.XfpYJQ.VLk2xpU18HqCIpu7EiZO7eekrE0')